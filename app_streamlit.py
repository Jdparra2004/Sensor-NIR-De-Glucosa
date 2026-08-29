"""
app_streamlit.py — Interfaz interactiva del Biosensor NIR
PROYECTO: Simulación paramétrica de detección óptica NIR de glucosa en sudor
"""

import io
import os
import time
import streamlit as st
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.modelo_optico import ModeloBeerLambertNIR, ModeloPLSRegresionNIR
from core.modelo_microfluido import ModeloMicrofluido

st.set_page_config(page_title="Biosensor NIR - Glucosa en Sudor", layout="wide")

# Configuración estándar de leyenda fija inferior
LEYENDA_INFERIOR = dict(
    orientation="h",
    yanchor="top",
    y=-0.25,
    xanchor="center",
    x=0.5
)

# --- ESTADO DE LA SESIÓN PARA PANTALLA DE BIENVENIDA ---
if 'show_info' not in st.session_state:
    st.session_state.show_info = True

def display_welcome_info():
    """Muestra la ventana de bienvenida con información del proyecto y disclaimer."""
    st.info("### Bienvenido al Biosensor NIR: Simulador Paramétrico")
    st.markdown("""
    Esta aplicación es una **herramienta de simulación numérica** diseñada para explorar los parámetros de diseño en sistemas de detección óptica de glucosa en sudor mediante espectroscopia NIR.

    **¿Qué puedes hacer aquí?**
    *   **Simulación Óptica:** Ajustar la longitud de onda y el camino óptico para analizar la absorbancia neta.
    *   **Análisis Microfluídico:** Evaluar el régimen de flujo (Reynolds) y el tiempo de residencia.
    *   **Sensibilidad:** Analizar cómo los cambios geométricos afectan la capacidad de detección.
    *   **Inferencia Clínica:** Procesar lotes de datos para estimar concentraciones de glucosa y clasificar resultados metabólicos.

    **IMPORTANTE - DISCLAIMER DE DISEÑO:**
    Este software es exclusivamente una **herramienta de simulación para diseño y exploración de parámetros**. 
    **NO** es un dispositivo médico, ni proporciona resultados clínicos, diagnósticos ni decisiones técnicas finales. Los resultados son proyecciones basadas en modelos teóricos (física-matemática) y deben utilizarse únicamente para evaluar la viabilidad de parámetros de diseño durante la fase de desarrollo.
    """)
    if st.button("Entendido y cerrar"):
        st.session_state.show_info = False
        st.rerun()

# Si debe mostrarse la información, la mostramos
if st.session_state.show_info:
    display_welcome_info()
    st.stop() # Detenemos ejecución para que solo se vea la info

# --- SIDEBAR ---
st.sidebar.header("Parámetros de Diseño y Simulación")

if st.sidebar.button("Ver Guía y Disclaimer"):
    st.session_state.show_info = True
    st.rerun()

st.sidebar.subheader("Óptica NIR")
lambda_nm = st.sidebar.slider("Longitud de onda (λ) [nm]", 1000, 1700, 1600, 1)
L_mm = st.sidebar.slider("Camino óptico (L) [mm]", 0.1, 2.0, 1.0, 0.1)
c_sim = st.sidebar.slider("Concentración de glucosa (C) [mM]", 0.01, 1.0, 0.20, 0.01)
noise_instrumental = st.sidebar.checkbox("Inyectar ruido fotométrico instrumental (±5%)", value=False)

st.sidebar.subheader("Microfluídica")
Q_nlmin = st.sidebar.slider("Caudal volumétrico (Q) [nL/min]", 1.0, 10.0, 5.0, 0.1)
w_um = st.sidebar.slider("Ancho del canal (w) [µm]", 50, 500, 200, 10)
h_um = st.sidebar.slider("Alto del canal (h) [µm]", 10, 200, 50, 10)
largo_mm = st.sidebar.slider("Largo de celda [mm]", 0.5, 5.0, 1.0, 0.5)

st.sidebar.subheader("Calibración Empírica")
alpha = st.sidebar.slider("Factor de escala (α)", 0.0, 1000.0, 1.0, 0.1)
beta = st.sidebar.slider("Sesgo (β)", -100.0, 100.0, 0.0, 0.1)

# --- MODELOS ---
modelo_optico = ModeloBeerLambertNIR(longitud_optica_mm=L_mm, alpha=alpha, beta=beta)
modelo_micro = ModeloMicrofluido(w_um, h_um, largo_mm, Q_nlmin)

def aplicar_ruido(valor):
    return valor * np.random.uniform(0.95, 1.05) if noise_instrumental else valor

def generar_excel_multihoja_estetico(df_resultado, lambda_val, L_val):
    """Genera un libro Excel multi-hoja con formato profesional, colores y estilos."""
    output = io.BytesIO()
    
    # 1. Escritura estructurada con Pandas
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Hoja 1: Resultados Detallados
        df_resultado.to_excel(writer, sheet_name='Resultados_Analisis', index=False)
        
        # Hoja 2: Resumen Fisiológico y Distribución
        if "Clasificación_Metabólica" in df_resultado.columns:
            conteo = df_resultado["Clasificación_Metabólica"].value_counts().reset_index()
            conteo.columns = ["Categoría Fisiológica", "Total Muestras"]
            conteo["Porcentaje (%)"] = (conteo["Total Muestras"] / len(df_resultado) * 100).round(2)
            conteo.to_excel(writer, sheet_name='Distribucion_Metabolica', index=False)
        
        # Hoja 3: Parámetros Ópticos y Métricas
        c_validos = df_resultado["Glucosa_Estimada_mM"].dropna()
        metricas = {
            "Parámetro de Simulación": [
                "Longitud de onda de análisis (λ)",
                "Camino óptico configurado (L)",
                "Total de muestras evaluadas",
                "Concentración media estimada",
                "Concentración mínima detectada",
                "Concentración máxima detectada"
            ],
            "Valor": [
                f"{lambda_val} nm",
                f"{L_val} mm",
                len(df_resultado),
                f"{c_validos.mean():.4f} mM" if not c_validos.empty else "N/A",
                f"{c_validos.min():.4f} mM" if not c_validos.empty else "N/A",
                f"{c_validos.max():.4f} mM" if not c_validos.empty else "N/A"
            ]
        }
        if "Error Relativo (%)" in df_resultado.columns and not df_resultado["Error Relativo (%)"].dropna().empty:
            metricas["Parámetro de Simulación"].append("Error relativo medio poblacional")
            metricas["Valor"].append(f"{df_resultado['Error Relativo (%)'].dropna().mean():.2f} %")
            
        df_params = pd.DataFrame(metricas)
        df_params.to_excel(writer, sheet_name='Parametros_Diseno', index=False)

    # 2. Post-procesamiento estético con OpenPyXL
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    
    # Estilos tipográficos y cromáticos
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    fill_normal = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_alerta = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_hiper = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    for ws in wb.worksheets:
        ws.sheet_state = 'visible'
        ws.views.sheetView[0].showGridLines = True
        
        # Formatear encabezado
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        # Formatear cuerpo y colorear categorías clínicas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = center_align
                
                # Resaltado condicional por texto de diagnóstico
                if str(cell.value) == "Normal":
                    cell.fill = fill_normal
                elif "Alerta" in str(cell.value):
                    cell.fill = fill_alerta
                elif "Elevado" in str(cell.value) or "Hiperglucemia" in str(cell.value):
                    cell.fill = fill_hiper

        # Autoajuste dinámico de ancho de columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.active = 0
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()

# --- PÁGINAS ---
st.title("Biosensor NIR: Simulación Integrada")
tab1, tab2, tab3, tab4 = st.tabs(["Óptica NIR", "Microfluídica", "Sensibilidad", "Inferencia Clínica"])

# Tab 1: Óptica
with tab1:
    with st.expander("Información del Análisis", expanded=False):
        st.markdown("Este módulo caracteriza la respuesta óptica del biosensor basándose en la Ley de Beer-Lambert. Calcula la absorbancia neta considerando el fenómeno de desplazamiento de agua, permitiendo visualizar la relación entre la concentración de glucosa y la absorbancia, así como el perfil espectral en la ventana de detección seleccionada.")
    
    c_range = np.linspace(0.01, 1.0, 100)
    abs_vals = [aplicar_ruido(modelo_optico.absorbancia(c, lambda_nm)) for c in c_range]
    
    col1, col2 = st.columns(2)
    fig1 = go.Figure()
    fig1.add_trace(go.Scatter(x=c_range, y=abs_vals, mode="lines", name="Absorbancia neta (Beer-Lambert corregido)"))
    fig1.update_layout(
        title="Absorbancia neta vs Concentración",
        xaxis_title="C [mM]",
        yaxis_title="A [u.a.]",
        height=380,
        showlegend=True,
        legend=LEYENDA_INFERIOR
    )
    col1.plotly_chart(fig1)

    lambdas, spect = modelo_optico.espectro_completo(c_sim, lambdas=np.linspace(1000, 1700, 100))
    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(x=lambdas, y=spect, mode="lines", name=f"Espectro NIR (C = {c_sim} mM)"))
    fig2.add_vrect(x0=1600, x1=1700, fillcolor="lightgray", opacity=0.3, annotation_text="Ventana 1600-1700 nm")
    fig2.update_layout(
        title="Espectro NIR",
        xaxis_title="λ [nm]",
        yaxis_title="A [u.a.]",
        height=380,
        showlegend=True,
        legend=LEYENDA_INFERIOR
    )
    col2.plotly_chart(fig2)

    with st.container(border=True):
        st.latex(r"C_{\text{final}} = \alpha \cdot \left( \frac{|A_{\text{neta}}|}{|\epsilon_{\text{g}}(\lambda) - \epsilon_{\text{w}}(\lambda) \cdot \delta_{\text{w}}| \cdot L} \right) + \beta")
        A_actual = modelo_optico.absorbancia(c_sim, lambda_nm)
        st.info(
            rf"Configuración: $\lambda = {lambda_nm}\text{{ nm}}$, $L = {L_mm}\text{{ mm}}$, "
            rf"$\alpha = {alpha}$, $\beta = {beta}$. "
            rf"Para una concentración teórica de $C = {c_sim}\text{{ mM}}$, "
            rf"la estimación ajustada resulta en **{modelo_optico.concentracion_inversa(A_actual, lambda_nm):.5f} mM**."
        )

# Tab 2: Microfluídica
with tab2:
    with st.expander("Información del Análisis", expanded=False):
        st.markdown("Este módulo evalúa las propiedades hidrodinámicas del fluido dentro del canal microfluídico. Calcula parámetros críticos para el diseño, incluyendo el número de Reynolds para verificar la laminaridad del flujo y el tiempo de residencia para determinar la interacción óptima fluido-sensor.")
    
    Q_range = np.linspace(1.0, 10.0, 50)
    Re_vals = [ModeloMicrofluido(w_um, h_um, largo_mm, q).numero_reynolds() for q in Q_range]
    tr_vals = [ModeloMicrofluido(w_um, h_um, largo_mm, q).tiempo_residencia_s() for q in Q_range]
    
    col1, col2 = st.columns(2)
    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=Q_range, y=Re_vals, mode="lines", name="Número de Reynolds (Re)"))
    fig3.add_trace(go.Scatter(x=[Q_range[0], Q_range[-1]], y=[1.0, 1.0], mode="lines", line=dict(dash="dash", color="red"), name="Límite laminar (Re = 1.0)"))
    fig3.update_layout(
        title="Número de Reynolds vs Caudal",
        xaxis_title="Q [nL/min]",
        yaxis_title="Re",
        height=380,
        showlegend=True,
        legend=LEYENDA_INFERIOR
    )
    col1.plotly_chart(fig3)
    
    fig4 = go.Figure()
    fig4.add_trace(go.Scatter(x=Q_range, y=tr_vals, mode="lines", name="Tiempo de residencia (t_r)"))
    fig4.update_layout(
        title="Tiempo de residencia vs Caudal",
        xaxis_title="Q [nL/min]",
        yaxis_title="t_r [s]",
        height=380,
        showlegend=True,
        legend=LEYENDA_INFERIOR
    )
    col2.plotly_chart(fig4)
    
    with st.container(border=True):
        st.latex(r"Re = \frac{2\rho Q}{\mu(w+h)} \quad \text{y} \quad t_r = \frac{V_{\text{celda}}}{Q}")
        re_actual = modelo_micro.numero_reynolds()
        st.info(
            rf"Con $Q = {Q_nlmin}\text{{ nL/min}}$, $Re = \mathbf{{{re_actual:.6f}}}$ "
            rf"({'Régimen Laminar' if re_actual < 1 else 'Flujo No Laminar'}). "
            rf"Velocidad media: **{modelo_micro.velocidad_media_m_s()*1e6:.2f} µm/s**. "
            rf"Tiempo de residencia: **{modelo_micro.tiempo_residencia_s():.2f} s**."
        )

# Tab 3: Sensibilidad
with tab3:
    with st.expander("Información del Análisis", expanded=False):
        st.markdown("Este estudio analiza cómo la longitud del camino óptico afecta la sensibilidad local (dA/dC) del biosensor. El objetivo es identificar configuraciones geométricas que maximicen la señal de detección sin degradar la selectividad del sistema.")
    
    L_range = np.linspace(0.1, 2.0, 50)
    sens_vals = [ModeloBeerLambertNIR(L).sensibilidad(lambda_nm) for L in L_range]
    abs_vals_L = [ModeloBeerLambertNIR(L).absorbancia(c_sim, lambda_nm) for L in L_range]
    
    col1, col2 = st.columns(2)
    fig5 = go.Figure()
    fig5.add_trace(go.Scatter(x=L_range, y=sens_vals, mode="lines", name="Sensibilidad local (dA/dC)"))
    fig5.update_layout(
        title="Sensibilidad vs Camino óptico",
        xaxis_title="L [mm]",
        yaxis_title="dA/dC [mM⁻¹]",
        height=380,
        showlegend=True,
        legend=LEYENDA_INFERIOR
    )
    col1.plotly_chart(fig5)
    
    fig6 = go.Figure()
    fig6.add_trace(go.Scatter(x=L_range, y=abs_vals_L, mode="lines", name=f"Absorbancia (C = {c_sim} mM)"))
    fig6.update_layout(
        title="Absorbancia vs Camino óptico",
        xaxis_title="L [mm]",
        yaxis_title="A [u.a.]",
        height=380,
        showlegend=True,
        legend=LEYENDA_INFERIOR
    )
    col2.plotly_chart(fig6)
    
    with st.container(border=True):
        st.latex(r"\text{Sensibilidad (física)} = \frac{\partial A}{\partial C} = \left(\epsilon_g(\lambda) - \epsilon_w(\lambda) \cdot \delta_w\right) \cdot L")
        st.info("Nota: Esta es la sensibilidad teórica del modelo físico Beer-Lambert. El modelo PLS-R multivariante utiliza un vector de pesos ($b_{PLS}$) entrenado sobre todo el espectro para maximizar la covarianza entre la absorbancia y la concentración.")

# Tab 4: Inferencia
with tab4:
    with st.expander("Información del Análisis", expanded=False):
        st.markdown("Motor de inferencia para la estimación de concentración de glucosa a partir de valores de absorbancia. Permite el análisis puntual o el procesamiento de lotes mediante carga de archivos CSV, clasificando las muestras según umbrales metabólicos fisiológicos.")

    st.subheader("Inferencia Clínica")
    A_med = st.number_input("Absorbancia medida (A)", value=-0.05, step=0.001, format="%.5f")
    if st.button("Estimar"):
        c_est = modelo_optico.concentracion_inversa(A_med, lambda_nm)
        st.write(f"Concentración estimada: **{c_est:.4f} mM** — Clasificación: **{modelo_optico.evaluar_clasificacion_fisiologica(c_est)}**")
        
    st.markdown("---")
    st.subheader("Procesamiento por Lotes")
    
    with st.expander("ℹ️ Guía de formato y Fundamentos Teóricos", expanded=False):
        st.markdown(rf"""
        **Condiciones de Análisis Activas:**
        * Los datos del archivo se evalúan en tiempo real con la **Longitud de onda ($\lambda = {lambda_nm}\text{{ nm}}$)** y el **Camino óptico ($L = {L_mm}\text{{ mm}}$)** configurados en la barra lateral (*sidebar*).
        
        **Flujo de Procesamiento Espectral (Multivariante):**
        1. **Pre-procesamiento:** Filtrado de 2956 canales para eliminar regiones ruidosas (<500nm, crossover 1090-1110nm, absorción de agua 1800-2100nm, >2300nm).
        2. **Calibración:** Entrenamiento PLS-R con selección automática de componentes latentes (CV-5fold).
        3. **Inferencia:** Cálculo de concentración mediante regresión sobre vectores de carga optimizados.
        
        **Formulación Matemática (PLS-R):**
        * **Descomposición:** $X = T P^T + E$, $y = T q + f$
        * **Regresión:** $b_{{PLS}} = W (P^T W)^{{-1}} q$
        * **Predicción:** $C_{{pred}} = b_0 + X_{{valid}} b_{{PLS}}$
        * **Ajuste:** $C_{{final}} = \alpha \cdot C_{{pred}} + \beta$

        **Estructura Requerida en el CSV/Parquet:**
        * **Columna de absorbancia (Legacy):** `absorbancia`, `A`, etc.
        * **Matriz Espectral (Multivariante):** Columnas con encabezados numéricos (ej. `1000.5`, `1001.0`).
        * **Columna de referencia (Opcional):** `Glucosa_Real_mM` (para métricas RMSEP).
        """)

    uploaded = st.file_uploader("Subir archivo de muestras (CSV, Parquet, TXT)", type=["csv", "parquet", "txt"])
    
    if uploaded is not None:
        progreso_contenedor = st.container()
        barra_progreso = progreso_contenedor.progress(0)
        estado_texto = progreso_contenedor.empty()
        
        try:
            estado_texto.text("Paso 1/4: Leyendo archivo en memoria...")
            barra_progreso.progress(25)
            
            # Detectar tipo de archivo
            file_ext = os.path.splitext(uploaded.name)[1].lower()
            
            if file_ext == '.parquet':
                df_lote = pd.read_parquet(uploaded)
            elif file_ext == '.txt':
                df_lote = pd.read_csv(uploaded, sep='\t', encoding='utf-16')
            else:
                try:
                    uploaded.seek(0)
                    df_lote = pd.read_csv(uploaded, encoding='utf-8')
                except Exception:
                    uploaded.seek(0)
                    df_lote = pd.read_csv(uploaded, sep=None, engine="python")
            
            # Validación y limitación a 1,000 filas
            total_filas_original = len(df_lote)
            if total_filas_original > 1000:
                df_lote = df_lote.head(1000).copy()
                st.warning(f"El archivo contiene {total_filas_original:,} filas. Para garantizar la fluidez de la interfaz, se procesan y analizan las primeras 1,000 muestras.")

            estado_texto.text("Paso 2/4: Identificando canal óptico o espectro completo...")
            barra_progreso.progress(50)
            time.sleep(0.05)
            
            # --- DETECCIÓN DE ESPECTRO COMPLETO (400 - 2500 nm) ---
            espectro_cols = [c for c in df_lote.columns if c.replace('.','',1).isdigit()]
            
            if len(espectro_cols) > 100: # Heurística para detectar matriz espectral
                estado_texto.text("Procesando con modelo multivariante PLS-R...")
                
                # Instanciar modelo PLS
                pls_model = ModeloPLSRegresionNIR(alpha=alpha, beta=beta)
                
                # Intentar buscar columna de referencia
                candidatos_ref = ['Glucose (mM)', 'glucosa_referencia_mM', 'Glucosa_Real_mM', 'glucosa_mM', 'glucose_mM', 'C_real']
                col_ref = next((c for c in candidatos_ref if c in df_lote.columns), None)
                
                if col_ref:
                    # Entrenamiento y predicción
                    y_series = pd.to_numeric(df_lote[col_ref], errors="coerce").fillna(0)
                    pls_model.entrenar_calibracion(df_lote, y_series)
                    df_lote["Glucosa_Estimada_mM"] = pls_model.predecir(df_lote, alpha=alpha, beta=beta)
                    
                    # Cálculo de RMSEP
                    mse = np.mean((df_lote["Glucosa_Estimada_mM"] - y_series)**2)
                    rmse = np.sqrt(mse)
                    st.write(f"**Métricas PLS-R:** RMSEP = {rmse:.4f} mM, Componentes óptimos = {pls_model.n_componentes_optimo}")
                else:
                    st.error("No se encontró columna de referencia para calibración del modelo PLS-R.")
            
            else:
                # --- LÓGICA UNIVARIANTE LEGACY ---
                col_abs = None
                candidatos_abs = ['absorbancia_1600nm', 'absorbancia_1650nm', 'absorbancia_medida', 'absorbancia', 'Absorbance', 'A']
                for cand in candidatos_abs:
                    if cand in df_lote.columns:
                        col_abs = cand
                        break
                
                if col_abs is None:
                    for col in df_lote.columns:
                        if 'abs' in col.lower():
                            col_abs = col
                            break

                if col_abs is not None:
                    estado_texto.text(rf"Paso 3/4: Ejecutando inferencia inversa ($\lambda={lambda_nm}\text{{ nm}}$, $L={L_mm}\text{{ mm}}$)...")
                    barra_progreso.progress(75)
                    
                    valores_abs = pd.to_numeric(df_lote[col_abs], errors="coerce")
                    df_lote["Glucosa_Estimada_mM"] = valores_abs.apply(
                        lambda a: modelo_optico.concentracion_inversa(float(a), lambda_nm, alpha=alpha, beta=beta) if pd.notna(a) else np.nan
                    ).round(4)
                    
                    candidatos_ref = ['glucosa_referencia_mM', 'Glucosa_Real_mM', 'glucosa_mM', 'glucose_mM', 'C_real']
                    col_ref = next((c for c in candidatos_ref if c in df_lote.columns), None)
            
            # --- CÁLCULO DE MÉTRICAS SI HAY REFERENCIA ---
            if 'Glucosa_Estimada_mM' in df_lote.columns:
                candidatos_ref = ['Glucose (mM)', 'glucosa_referencia_mM', 'Glucosa_Real_mM', 'glucosa_mM', 'glucose_mM', 'C_real']
                col_ref = next((c for c in candidatos_ref if c in df_lote.columns), None)
                if col_ref is not None:
                    c_real = pd.to_numeric(df_lote[col_ref], errors="coerce")
                    # Renombrado de Error_% a Error Relativo (%)
                    df_lote["Error Relativo (%)"] = (np.abs(df_lote["Glucosa_Estimada_mM"] - c_real) / np.where(c_real != 0, c_real, 1e-12) * 100).round(2)
                
                df_lote["Clasificación_Metabólica"] = df_lote["Glucosa_Estimada_mM"].apply(
                    lambda c: modelo_optico.evaluar_clasificacion_fisiologica(c) if pd.notna(c) else "Indeterminado"
                )
                
                estado_texto.text("Paso 4/4: Consolidando resultados y libro de reporte...")
                barra_progreso.progress(100)
                time.sleep(0.05)
                
                barra_progreso.empty()
                estado_texto.success(f"Procesamiento completado: {len(df_lote):,} muestras analizadas bajo λ = {lambda_nm} nm y L = {L_mm} mm.")
                
                st.dataframe(df_lote)
                
                # --- GRÁFICOS DE ANÁLISIS DEL LOTE ---
                st.markdown("#### Análisis Estadístico y Fisiológico del Lote")
                col_g1, col_g2 = st.columns(2)
                
                # Gráfico 1: Conteo por Categoría Fisiológica
                conteo_df = df_lote["Clasificación_Metabólica"].value_counts().reset_index()
                conteo_df.columns = ["Categoría", "Muestras"]
                
                colores_map = {
                    "Normal": "#2ca02c",
                    "Rango de Alerta / Sospecha de Prediabetes": "#ff7f0e",
                    "Nivel Elevado / Sospecha Hiperglucemia": "#d62728",
                    "Fuera de rango analítico / Indetectable": "#7f7f7f"
                }
                bar_colors = [colores_map.get(cat, "#1f77b4") for cat in conteo_df["Categoría"]]
                
                fig_lote_cat = go.Figure()
                fig_lote_cat.add_trace(go.Bar(
                    x=conteo_df["Categoría"],
                    y=conteo_df["Muestras"],
                    marker_color=bar_colors,
                    name="Muestras por Estado"
                ))
                fig_lote_cat.update_layout(
                    title="<b>Distribución de Categorías Fisiológicas</b>",
                    xaxis_title="Estado Metabólico",
                    yaxis_title="Cantidad de Muestras",
                    height=350,
                    showlegend=False,
                    margin=dict(l=40, r=40, t=50, b=40)
                )
                col_g1.plotly_chart(fig_lote_cat)
                
                # Gráfico 2: Dispersión de Concentración Estimada con Umbrales
                fig_lote_disp = go.Figure()
                indices_muestras = list(range(1, len(df_lote) + 1))
                
                fig_lote_disp.add_trace(go.Scatter(
                    x=indices_muestras,
                    y=df_lote["Glucosa_Estimada_mM"],
                    mode="markers",
                    name="Glucosa Estimada (mM)",
                    marker=dict(color="#1f77b4", size=5, opacity=0.7)
                ))
                fig_lote_disp.add_hline(y=0.20, line_dash="dash", line_color="green", annotation_text="Límite Normal (0.20 mM)")
                fig_lote_disp.add_hline(y=0.40, line_dash="dash", line_color="red", annotation_text="Umbral Hiperglucemia (0.40 mM)")
                
                fig_lote_disp.update_layout(
                    title="<b>Concentración Estimada por Muestra</b>",
                    xaxis_title="Índice de Muestra",
                    yaxis_title="Glucosa [mM]",
                    height=350,
                    showlegend=True,
                    legend=LEYENDA_INFERIOR,
                    margin=dict(l=40, r=40, t=50, b=40)
                )
                col_g2.plotly_chart(fig_lote_disp)
                
                # Exportación
                col_btn1, col_btn2 = st.columns(2)
                excel_bytes = generar_excel_multihoja_estetico(df_lote, lambda_nm, L_mm)
                
                with col_btn1:
                    if excel_bytes is not None:
                        st.download_button(
                            label="Descargar Reporte Completo en Excel (.xlsx)",
                            data=excel_bytes,
                            file_name="Reporte_Biosensor_NIR.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.info("Exportación directa a CSV disponible.")
                        
                with col_btn2:
                    st.download_button(
                        label="Descargar en formato CSV",
                        data=df_lote.to_csv(index=False).encode("utf-8"),
                        file_name="resultados_procesamiento_lote.csv",
                        mime="text/csv"
                    )
            else:
                barra_progreso.empty()
                estado_texto.error("No se detectó una columna de absorbancia válida en el archivo cargado.")
        except Exception as e:
            barra_progreso.empty()
            estado_texto.error(f"Error al procesar el archivo: {e}")